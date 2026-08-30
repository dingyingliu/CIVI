"""Export verified QA pairs to Excel.

Produces a single ``.xlsx`` file with columns: question_id, question,
answer, category, institution, city, source_url, page_title,
supporting_quotes, model_used.  MC questions include formatted options
in the question column.

If any QA pair carries a ``_cell`` value (set by the generator from the
3-country COFOG config), the workbook is split into one sheet per
COFOG cell — ``03 — Public Order and Safety``, ``04 — Economic Affairs``,
``07 — Health``, ``10 — Social Protection``.  (The ``Cell`` prefix is
dropped so the longest name fits Excel's 31-char sheet-name limit;
``Cell 03 — Public Order and Safety`` would be 33.)  Within each sheet,
rows are sorted by ``(institution_short_name, label)``.  Otherwise (no
COFOG metadata available) a single ``QA Pairs`` sheet is written.
"""

from __future__ import annotations

import logging
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from src.generate.question_builder_one_call import QAPair

logger = logging.getLogger(__name__)


CELL_SHEET_NAMES: dict[str, str] = {
    "03": "03 — Public Order and Safety",
    "04": "04 — Economic Affairs",
    "07": "07 — Health",
    "10": "10 — Social Protection",
}

HEADERS: list[str] = [
    "question_id", "question", "answer",
    "category", "institution", "city",
    "source_url", "page_title",
    "supporting_quotes", "model_used",
]

COL_WIDTH_CAPS: dict[str, int] = {
    "source_url": 60,
    "page_title": 50,
    "supporting_quotes": 80,
}


class QAExporter:
    """Export QA pairs to an Excel workbook.

    Attributes:
        output_path: Path where the ``.xlsx`` file is written.
        city: Fallback city label written when a row has no
            per-row ``_institution_short_name`` set on the QAPair.
    """

    def __init__(
        self,
        output_path: str | Path = "output/qa_pairs.xlsx",
        city: str = "",
    ) -> None:
        self.output_path = Path(output_path)
        self.city = city

    def export(self, qa_pairs: list[QAPair]) -> Path:
        """Write all QA pairs to an Excel file.

        If any pair has a ``_cell`` value, splits into per-cell sheets
        (sorted by short_name then label). Otherwise writes a single
        ``QA Pairs`` sheet.
        """
        self.output_path.parent.mkdir(parents=True, exist_ok=True)

        wb = Workbook()
        wb.remove(wb.active)  # start clean — sheets are created per-cell below

        if any(getattr(qa, "_cell", "") for qa in qa_pairs):
            self._write_cell_sheets(wb, qa_pairs)
        else:
            self._write_single_sheet(wb, qa_pairs)

        wb.save(self.output_path)
        logger.info("Exported %d QA pairs to %s", len(qa_pairs), self.output_path)
        return self.output_path

    def _write_cell_sheets(self, wb: Workbook, qa_pairs: list[QAPair]) -> None:
        """Group by ``_cell`` and write one sheet per known cell code."""
        by_cell: dict[str, list[QAPair]] = {}
        for qa in qa_pairs:
            cell = getattr(qa, "_cell", "") or "other"
            by_cell.setdefault(cell, []).append(qa)

        # Known cells first in canonical order, then any unexpected cells
        # in lexicographic order so the workbook is deterministic.
        ordered_cells = [c for c in CELL_SHEET_NAMES if c in by_cell]
        ordered_cells += sorted(c for c in by_cell if c not in CELL_SHEET_NAMES)

        for cell in ordered_cells:
            sheet_name = CELL_SHEET_NAMES.get(cell, f"Cell {cell}")
            ws = wb.create_sheet(title=sheet_name)
            sorted_qas = sorted(by_cell[cell], key=self._sort_key)
            self._write_sheet(ws, sorted_qas)

    def _write_single_sheet(self, wb: Workbook, qa_pairs: list[QAPair]) -> None:
        ws = wb.create_sheet(title="QA Pairs")
        self._write_sheet(ws, qa_pairs)

    @staticmethod
    def _sort_key(qa: QAPair) -> tuple[str, str]:
        return (
            getattr(qa, "_institution_short_name", "") or "",
            getattr(qa, "_label", "") or "",
        )

    def _write_sheet(self, ws, qa_pairs: list[QAPair]) -> None:
        ws.append(HEADERS)
        for cell in ws[1]:
            cell.font = Font(bold=True)

        for qa in qa_pairs:
            question_text = self._format_question(qa)
            source_url = (qa.citation or {}).get("source_uri", "")
            page_title = (qa.citation or {}).get("title", "")
            quotes_text = self._format_quotes(qa)
            institution = getattr(qa, "_institution", "") or ""
            short_name = getattr(qa, "_institution_short_name", "") or ""
            # Per-row short_name overrides the constructor's city when present
            # (it's the right granularity for unified-shape data).
            city_value = short_name or self.city

            ws.append([
                qa.question_id,
                question_text,
                qa.answer,
                qa.category,
                institution,
                city_value,
                source_url,
                page_title,
                quotes_text,
                qa.model_used,
            ])

        for col in ws.columns:
            header = col[0].value or ""
            cap = COL_WIDTH_CAPS.get(header, 80)
            max_len = 0
            for cell in col:
                if cell.value:
                    lines = str(cell.value).split("\n")
                    max_len = max(max_len, max(len(line) for line in lines))
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, cap)

    def _format_question(self, qa: QAPair) -> str:
        if qa.options:
            options_str = "\n".join(
                f"  {opt['label']}. {opt['text']}" for opt in qa.options
            )
            return f"{qa.question}\n\n{options_str}"
        return qa.question

    def _format_quotes(self, qa: QAPair) -> str:
        if not qa.quotes:
            return ""
        return "\n".join(f'"{q}"' for q in qa.quotes)
