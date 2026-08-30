"""Per-model stratified source-injection diagnostic.

Pick ``--n`` questions a single model got wrong, stratified across 9
jurisdictions × 4 COFOG categories. For each, re-run the model with the
source page text injected into the prompt and tools disabled, then
classify the failure.

Read-only against the 9 benchmark Excels and corpus DBs. Writes ONE new
diagnostic Excel; does not mutate any existing benchmark file.

Usage:
    uv run python scripts/run_diagnostic_per_model.py \
        --model qwen-3.6-plus \
        --n 100 \
        --workers 10 \
        --output output/diagnostic_results/qwen_diag_n100.xlsx
"""
from __future__ import annotations

import argparse
import logging
import os
import random
import re
import sys
import time
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path

import openpyxl
import pandas as pd
from openpyxl.styles import Font

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

# Minimal .env loader
_env = ROOT / ".env"
if _env.exists():
    for line in _env.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        k, v = line.split("=", 1)
        os.environ.setdefault(k.strip(), v.strip().strip("'\""))

from src.benchmark.answer_extractor import GeminiAnswerExtractor  # noqa: E402
from src.benchmark.benchmark_text_qa import (  # noqa: E402
    DEFAULT_EVAL_MODELS,
    TextQABenchmark,
    model_extras,
)
from src.diagnostic.source_loader import SourceLoader, normalize_url  # noqa: E402
from src.llm.openrouter_client import OpenRouterClient  # noqa: E402

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s %(name)s: %(message)s",
)
logger = logging.getLogger("diagnostic")

# 9 benchmark Excels and their matching corpus DB short-name keys.
JURISDICTIONS: list[tuple[str, Path, str]] = [
    ("AU federal",     ROOT / "output/3country_benchmark_results/au_federal_full_results.xlsx",     "australia_federal"),
    ("AU state",       ROOT / "output/3country_benchmark_results/au_state_full_results.xlsx",       "australia_state"),
    ("AU municipal",   ROOT / "output/3country_benchmark_results/au_municipal_full_results.xlsx",   "australia_municipal"),
    ("CA federal",     ROOT / "output/3country_benchmark_results/ca_federal_full_results.xlsx",     "canada_federal"),
    ("CA provincial",  ROOT / "output/3country_benchmark_results/ca_provincial_full_results.xlsx",  "canada_provincial"),
    ("CA municipal",   ROOT / "output/3country_benchmark_results/ca_municipal_full_results.xlsx",   "canada_municipal"),
    ("US federal",     ROOT / "output/3country_benchmark_results/us_federal_full_results.xlsx",     "united_states_federal"),
    ("US state",       ROOT / "output/3country_benchmark_results/us_state_full_results.xlsx",       "united_states_state"),
    ("US municipal",   ROOT / "output/3country_benchmark_results/us_municipal_full_results.xlsx",   "united_states_municipal"),
]

PROMPT_TEMPLATE = (
    "Use ONLY the source text below to answer. Do not use any outside knowledge.\n\n"
    "Source page:\n"
    "{source_text}\n\n"
    "Question:\n"
    "{question}\n"
)


# ───── Helpers ────────────────────────────────────────────────────────────


def truthy(v) -> bool:
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float)):
        if pd.isna(v):
            return False
        return bool(v)
    if isinstance(v, str):
        return v.strip().lower() in ("true", "1", "yes")
    return False


def is_failure(v) -> bool:
    """Row is a failure (correct=0). Treat NaN / missing / non-numeric as not-failed."""
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return False
    try:
        return int(v) == 0
    except (TypeError, ValueError):
        return False


def parse_cited_urls(raw) -> list[str]:
    if not isinstance(raw, str):
        return []
    s = raw.strip()
    if not s:
        return []
    return [u.strip() for u in s.split(" | ") if u.strip()]


# ───── Stratified sampler ─────────────────────────────────────────────────


def stratified_sample(
    pool_by_bucket: dict[tuple[str, str], list[dict]],
    n: int,
    seed: int = 42,
) -> list[dict]:
    """Pick ``n`` items spread evenly across buckets, with deficit redistribution.

    Buckets with fewer candidates than their target take all; the slack is
    redistributed evenly across buckets that still have headroom. Repeats
    until either ``n`` is reached or no bucket can yield more.
    """
    rng = random.Random(seed)
    buckets = {k: list(v) for k, v in pool_by_bucket.items()}
    n_buckets = len(buckets)
    if n_buckets == 0:
        return []

    chosen: list[dict] = []
    chosen_keys: set[tuple[str, int]] = set()  # (jurisdiction, row_idx) dedup

    # iterative top-up
    while len(chosen) < n:
        remaining_n = n - len(chosen)
        active = [k for k, v in buckets.items() if v]
        if not active:
            break
        per_bucket = max(1, remaining_n // len(active))
        for k in active:
            if len(chosen) >= n:
                break
            take = min(per_bucket, len(buckets[k]), n - len(chosen))
            picks = rng.sample(buckets[k], take)
            for p in picks:
                key = (p["jurisdiction"], p["row_idx"])
                if key in chosen_keys:
                    continue
                chosen.append(p)
                chosen_keys.add(key)
                buckets[k].remove(p)
        # If per_bucket==1 and we still have headroom, the second pass picks 1 more from each
    return chosen[:n]


# ───── Candidate pool builder ─────────────────────────────────────────────


def build_candidate_pool(
    model: str,
    loader: SourceLoader,
    exclude_qids: set[tuple[str, str]] | None = None,
) -> list[dict]:
    """For each row across all 9 Excels where the model was wrong, look up
    the source page text and add it to the candidate pool.

    If ``exclude_qids`` is given (set of ``(jurisdiction, question_id)``
    tuples), those rows are skipped — used to dedupe against questions
    already in an existing diagnostic Excel.
    """
    exclude_qids = exclude_qids or set()
    out: list[dict] = []
    for juris_label, xlsx_path, db_key in JURISDICTIONS:
        if not xlsx_path.exists():
            logger.warning("Missing Excel: %s", xlsx_path)
            continue
        df = pd.read_excel(xlsx_path, sheet_name="Results")
        correct_col = f"correct_{model}"
        resp_col = f"response_{model}"
        searched_col = f"searched_{model}"
        cited_col = f"cited_urls_{model}"
        if correct_col not in df.columns:
            logger.warning("%s missing column %s", juris_label, correct_col)
            continue

        n_wrong = 0
        n_kept = 0
        n_already_done = 0
        for row_idx, row in df.iterrows():
            if not is_failure(row.get(correct_col)):
                continue
            n_wrong += 1
            qid = str(row.get("question_id") or "").strip()
            if (juris_label, qid) in exclude_qids:
                n_already_done += 1
                continue
            source_url = str(row.get("source_url") or "").strip()
            if not source_url:
                continue
            city = str(row.get("city") or "").strip()
            text = loader.lookup(source_url, city)
            if not text:
                continue
            wc = len(text.split())
            n_kept += 1
            out.append({
                "jurisdiction":      juris_label,
                "xlsx_path":         str(xlsx_path),
                "row_idx":           int(row_idx),
                "question_id":       str(row.get("question_id") or "").strip(),
                "question_text":     str(row.get("question_text") or "").strip(),
                "gold_answer":       str(row.get("gold_answer") or "").strip(),
                "category":          str(row.get("category") or "").strip(),
                "city":              city,
                "source_url":        source_url,
                "source_word_count": wc,
                "source_text":       text,
                "original_response": str(row.get(resp_col) or "").strip(),
                "original_searched": truthy(row.get(searched_col)),
                "original_cited_urls": parse_cited_urls(row.get(cited_col)),
            })
        logger.info(
            "  %s: %d wrong | %d already done | %d kept",
            juris_label, n_wrong, n_already_done, n_kept,
        )
    return out


# ───── Diagnostic execution ───────────────────────────────────────────────


def diagnose_one(
    client: OpenRouterClient,
    extractor: GeminiAnswerExtractor,
    model_id: str,
    model_name: str,
    item: dict,
) -> dict:
    prompt = PROMPT_TEMPLATE.format(
        source_text=item["source_text"], question=item["question_text"],
    )
    extra_body, max_tokens = model_extras(model_id, "MC")
    try:
        result = client.chat_eval(
            model=model_id,
            messages=[{"role": "user", "content": prompt}],
            temperature=0,
            max_tokens=max_tokens,
            stage="diagnostic",
            tools=None,  # web search OFF
            extra_body=extra_body,
        )
    except Exception as exc:  # noqa: BLE001
        logger.warning("API failure for qid=%s: %s", item["question_id"], exc)
        return {**item, "diagnostic_response": "", "diagnostic_correct": 0,
                "failure_mode": "api_error"}

    response_text = result.get("text", "") or ""

    # Score with the canonical Flash extractor (same as the benchmark uses)
    valid_labels = sorted(TextQABenchmark._extract_valid_labels(item["question_text"]))
    gold = set(re.findall(r"[A-Z]", item["gold_answer"].upper())) & set(valid_labels)
    extracted = extractor.extract(
        response_text=response_text,
        valid_labels=valid_labels,
        qid=str(item["question_id"]),
        model_name=model_name,
        cache_key_extra="|diag",
    )
    pred = set(extracted)
    diag_correct = pred == gold

    # Classify failure mode
    if not diag_correct:
        failure_mode = "comprehension"
    elif not item["original_searched"]:
        failure_mode = "search_bypass"
    else:
        target = normalize_url(item["source_url"])
        cited = {normalize_url(u) for u in item["original_cited_urls"]}
        failure_mode = "grounding" if target and target in cited else "retrieval"

    return {
        **item,
        "diagnostic_response": response_text,
        "diagnostic_correct": "correct" if diag_correct else "incorrect",
        "failure_mode": failure_mode,
    }


# ───── Output writer ──────────────────────────────────────────────────────


def load_existing_qids(path: Path) -> set[tuple[str, str]]:
    """Read (jurisdiction, question_id) pairs already present in the
    Results sheet of an existing diagnostic Excel."""
    if not path.exists():
        return set()
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb["Results"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return set()
    hdrs = list(rows[0])
    if "question_id" not in hdrs or "jurisdiction" not in hdrs:
        return set()
    qid_i = hdrs.index("question_id")
    juris_i = hdrs.index("jurisdiction")
    return {(str(r[juris_i] or ""), str(r[qid_i] or "")) for r in rows[1:]}


def append_to_diagnostic_excel(
    output_path: Path, results: list[dict], model: str,
) -> None:
    """Append new rows under the existing Results sheet, then refresh the
    Summary sheet by re-aggregating across all rows."""
    wb = openpyxl.load_workbook(output_path)
    ws = wb["Results"]
    hdrs = [c.value for c in ws[1]]

    def col(name):
        return hdrs.index(name) + 1

    start_row = ws.max_row + 1
    for r_i, r in enumerate(results, start_row):
        ws.cell(row=r_i, column=col("question_id"),
                value=r["question_id"])
        ws.cell(row=r_i, column=col("jurisdiction"),
                value=r["jurisdiction"])
        ws.cell(row=r_i, column=col("category"), value=r["category"])
        ws.cell(row=r_i, column=col("city"),     value=r["city"])
        ws.cell(row=r_i, column=col("source_url"),
                value=r["source_url"])
        ws.cell(row=r_i, column=col("source_word_count"),
                value=r["source_word_count"])
        ws.cell(row=r_i, column=col("question_text"),
                value=r["question_text"])
        ws.cell(row=r_i, column=col("gold_answer"),
                value=r["gold_answer"])
        ws.cell(row=r_i, column=col(f"original_response_{model}"),
                value=r.get("original_response", ""))
        ws.cell(row=r_i, column=col(f"original_searched_{model}"),
                value=bool(r.get("original_searched")))
        ws.cell(row=r_i, column=col(f"original_cited_urls_{model}"),
                value=" | ".join(r.get("original_cited_urls", []) or []))
        ws.cell(row=r_i, column=col(f"diagnostic_response_{model}"),
                value=r.get("diagnostic_response", ""))
        ws.cell(row=r_i, column=col(f"diagnostic_correct_{model}"),
                value=r.get("diagnostic_correct", "incorrect"))
        ws.cell(row=r_i, column=col(f"failure_mode_{model}"),
                value=r.get("failure_mode", ""))

    # Refresh Summary by re-aggregating all rows
    fmode_i = col(f"failure_mode_{model}")
    diag_corr_i = col(f"diagnostic_correct_{model}")
    from collections import Counter
    modes: Counter = Counter()
    correct = total = 0
    for row in ws.iter_rows(min_row=2):
        modes[row[fmode_i - 1].value] += 1
        if row[diag_corr_i - 1].value == "correct":
            correct += 1
        total += 1

    s = wb["Summary"]
    for r in range(s.max_row, 1, -1):
        for c in range(1, s.max_column + 1):
            s.cell(row=r, column=c).value = None
    rows_out = [
        ("total_diagnosed", total),
        ("diagnostic_correct", correct),
    ]
    for mode in ["search_bypass", "retrieval", "grounding",
                 "comprehension", "api_error"]:
        rows_out.append((f"failure_mode={mode}", modes.get(mode, 0)))
    for r_i, (k, v) in enumerate(rows_out, 2):
        s.cell(row=r_i, column=1, value=k)
        s.cell(row=r_i, column=2, value=v)

    wb.save(output_path)
    logger.info("Appended %d rows to %s (total %d)", len(results),
                output_path, total)


def write_diagnostic_excel(results: list[dict], output_path: Path, model: str) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Results"
    headers = [
        "question_id", "jurisdiction", "category", "city",
        "source_url", "source_word_count",
        "question_text", "gold_answer",
        f"original_response_{model}",
        f"original_searched_{model}",
        f"original_cited_urls_{model}",
        f"diagnostic_response_{model}",
        f"diagnostic_correct_{model}",
        f"failure_mode_{model}",
    ]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = Font(bold=True)

    for r_i, r in enumerate(results, 2):
        ws.cell(row=r_i, column=1,  value=r["question_id"])
        ws.cell(row=r_i, column=2,  value=r["jurisdiction"])
        ws.cell(row=r_i, column=3,  value=r["category"])
        ws.cell(row=r_i, column=4,  value=r["city"])
        ws.cell(row=r_i, column=5,  value=r["source_url"])
        ws.cell(row=r_i, column=6,  value=r["source_word_count"])
        ws.cell(row=r_i, column=7,  value=r["question_text"])
        ws.cell(row=r_i, column=8,  value=r["gold_answer"])
        ws.cell(row=r_i, column=9,  value=r.get("original_response", ""))
        ws.cell(row=r_i, column=10, value=bool(r.get("original_searched")))
        ws.cell(row=r_i, column=11, value=" | ".join(r.get("original_cited_urls", []) or []))
        ws.cell(row=r_i, column=12, value=r.get("diagnostic_response", ""))
        ws.cell(row=r_i, column=13, value=r.get("diagnostic_correct", "incorrect"))
        ws.cell(row=r_i, column=14, value=r.get("failure_mode", ""))

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    ws2.cell(row=1, column=1, value="metric").font = Font(bold=True)
    ws2.cell(row=1, column=2, value="count").font = Font(bold=True)
    rows = [
        ("total_diagnosed", len(results)),
        ("diagnostic_correct", sum(
            1 for r in results if r.get("diagnostic_correct") == "correct"
        )),
    ]
    for mode in ["search_bypass", "retrieval", "grounding", "comprehension", "api_error"]:
        rows.append((f"failure_mode={mode}",
                     sum(1 for r in results if r.get("failure_mode") == mode)))
    for r_i, (k, v) in enumerate(rows, 2):
        ws2.cell(row=r_i, column=1, value=k)
        ws2.cell(row=r_i, column=2, value=v)

    wb.save(output_path)
    logger.info("Wrote %s", output_path)


# ───── Main ───────────────────────────────────────────────────────────────


def main() -> int:
    p = argparse.ArgumentParser()
    p.add_argument("--model", default="qwen-3.6-plus",
                   help="Model short name (must be in DEFAULT_EVAL_MODELS)")
    p.add_argument("--n", type=int, default=100,
                   help="Total questions to sample across 36 buckets")
    p.add_argument("--workers", type=int, default=10)
    p.add_argument("--seed", type=int, default=42)
    p.add_argument("--output", type=str,
                   default="output/diagnostic_results/qwen_diag_n100.xlsx")
    p.add_argument("--append-to", type=str, default=None,
                   help="Append results to an existing diagnostic Excel; "
                        "skips question_ids already present. Overrides --output.")
    args = p.parse_args()

    model = args.model
    if model not in DEFAULT_EVAL_MODELS:
        logger.error("Unknown model %r. Known: %s", model, list(DEFAULT_EVAL_MODELS.keys()))
        return 1
    model_id = DEFAULT_EVAL_MODELS[model]

    append_mode = bool(args.append_to)
    output_path = ROOT / (args.append_to or args.output)

    logger.info("=== Per-model diagnostic ===")
    logger.info("  model:     %s (%s)", model, model_id)
    logger.info("  target n:  %d", args.n)
    logger.info("  workers:   %d", args.workers)
    logger.info("  output:    %s", output_path)
    logger.info("  mode:      %s", "APPEND" if append_mode else "FRESH WRITE")

    exclude_qids: set[tuple[str, str]] = set()
    if append_mode:
        exclude_qids = load_existing_qids(output_path)
        logger.info("  existing qids in target: %d", len(exclude_qids))

    # 1. Build candidate pool
    logger.info("\n=== Stage 1: Build candidate pool ===")
    loader = SourceLoader(ROOT / "data")
    candidates = build_candidate_pool(
        model, loader, exclude_qids=exclude_qids,
    )
    logger.info("Total eligible candidates (after dedup): %d", len(candidates))

    # 2. Stratify
    logger.info("\n=== Stage 2: Stratify and sample %d ===", args.n)
    pool_by_bucket: dict[tuple[str, str], list[dict]] = defaultdict(list)
    for c in candidates:
        pool_by_bucket[(c["jurisdiction"], c["category"])].append(c)
    logger.info("Bucket counts (sorted):")
    for (j, cat), items in sorted(pool_by_bucket.items()):
        logger.info("  %-16s | %-26s | %d candidates", j, cat, len(items))
    sampled = stratified_sample(pool_by_bucket, args.n, seed=args.seed)
    logger.info("Sampled %d questions", len(sampled))

    # Print per-bucket sample distribution
    sample_counts: dict[tuple[str, str], int] = defaultdict(int)
    for s in sampled:
        sample_counts[(s["jurisdiction"], s["category"])] += 1
    logger.info("Sample distribution:")
    for (j, cat), n in sorted(sample_counts.items()):
        logger.info("  %-16s | %-26s | %d sampled", j, cat, n)

    # 3. Diagnose
    logger.info("\n=== Stage 3: Run %d diagnostic calls (%d workers) ===",
                len(sampled), args.workers)
    client = OpenRouterClient()
    extractor = GeminiAnswerExtractor()
    t0 = time.time()
    results: list[dict] = []
    with ThreadPoolExecutor(max_workers=args.workers) as pool:
        futures = {pool.submit(diagnose_one, client, extractor, model_id, model, item): item
                   for item in sampled}
        done = 0
        for fut in as_completed(futures):
            results.append(fut.result())
            done += 1
            if done % 10 == 0 or done == len(futures):
                logger.info("  progress: %d / %d (%.1fs elapsed)",
                            done, len(futures), time.time() - t0)
    extractor.flush_cache()
    logger.info("Diagnostic phase done in %.1fs", time.time() - t0)

    # 4. Write Excel
    if append_mode and output_path.exists():
        logger.info("\n=== Stage 4: Write Excel (APPEND) ===")
        append_to_diagnostic_excel(output_path, results, model)
    else:
        logger.info("\n=== Stage 4: Write Excel (FRESH) ===")
        write_diagnostic_excel(results, output_path, model)

    # 5. Summary
    correct = sum(
        1 for r in results if r.get("diagnostic_correct") == "correct"
    )
    logger.info("\n=== Summary ===")
    logger.info("  total diagnosed:    %d", len(results))
    logger.info("  diagnostic correct: %d (%.1f%%)",
                correct, 100 * correct / max(len(results), 1))
    logger.info("  failure modes:")
    by_mode: dict[str, int] = defaultdict(int)
    for r in results:
        by_mode[r.get("failure_mode", "")] += 1
    for mode in ("search_bypass", "retrieval", "grounding",
                 "comprehension", "api_error"):
        logger.info("    %-15s %d", mode, by_mode.get(mode, 0))

    return 0


if __name__ == "__main__":
    sys.exit(main())
