"""Flatten a multi-sheet jurisdiction QA xlsx into a single-sheet file.

The benchmark pipeline reads only the active sheet via ``pd.read_excel``.
Jurisdiction QA files have 4 COFOG sheets — this script concatenates them
into a single-sheet xlsx that the benchmark consumes unchanged. The
``category`` column is preserved so downstream analysis can group by COFOG.

Usage:
    uv run python scripts/concatenate_jurisdiction_sheets.py \\
        --input  output/3country_qa_pairs_final/australia_federal_qa_pairs.xlsx \\
        --output output/3country_benchmark_results/_au_federal_concat.xlsx
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font

REQUIRED_COLS = [
    "question_id",
    "question",
    "answer",
    "category",
    "institution",
    "city",
    "source_url",
    "page_title",
    "supporting_quotes",
]


def main(argv: list[str] | None = None) -> int:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("--input", type=Path, required=True,
                   help="Multi-sheet QA xlsx (e.g. <jurisdiction>_qa_pairs.xlsx)")
    p.add_argument("--output", type=Path, required=True,
                   help="Single-sheet flattened xlsx output path")
    args = p.parse_args(argv)

    if not args.input.exists():
        print(f"ERROR: input file not found: {args.input}", file=sys.stderr)
        return 2

    print(f"Input  : {args.input}")
    print(f"Output : {args.output}\n")

    wb_in = openpyxl.load_workbook(args.input, read_only=True, data_only=True)
    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = "All Categories"

    headers: list[str] | None = None
    per_sheet: list[tuple[str, int]] = []
    total = 0

    for sn in wb_in.sheetnames:
        ws = wb_in[sn]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        sheet_header = [str(h) if h is not None else "" for h in rows[0]]
        data = [r for r in rows[1:] if r is not None and not all(v is None for v in r)]

        if headers is None:
            headers = sheet_header
            ws_out.append(headers)
            for cell in ws_out[1]:
                cell.font = Font(bold=True)
        elif sheet_header != headers:
            print(f"  WARNING: header mismatch on sheet {sn!r}", file=sys.stderr)
            print(f"    first sheet: {headers}", file=sys.stderr)
            print(f"    this sheet:  {sheet_header}", file=sys.stderr)
            return 3

        for r in data:
            ws_out.append(list(r))
        per_sheet.append((sn, len(data)))
        total += len(data)

    wb_in.close()

    # Sanity check: required columns present
    if headers is None:
        print("ERROR: no sheets had a header row", file=sys.stderr)
        return 4
    missing = [c for c in REQUIRED_COLS if c not in headers]
    if missing:
        print(f"ERROR: input is missing required columns: {missing}", file=sys.stderr)
        print(f"  headers found: {headers}", file=sys.stderr)
        return 5

    # Light formatting on output
    for col in ws_out.columns:
        col_letter = col[0].column_letter
        max_len = 0
        for cell in col:
            if cell.value is None:
                continue
            for line in str(cell.value).split("\n"):
                if len(line) > max_len:
                    max_len = len(line)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws_out.column_dimensions[col_letter].width = min(max_len + 2, 80)

    args.output.parent.mkdir(parents=True, exist_ok=True)
    wb_out.save(args.output)

    # Report
    print(f"Per-sheet input row counts:")
    for sn, n in per_sheet:
        print(f"  {sn}: {n} rows")
    print(f"\nOutput total: {total} rows")
    print(f"Sanity check: all {len(REQUIRED_COLS)} required base columns present [OK]")
    print(f"  required: {REQUIRED_COLS}")
    print(f"\nWrote {args.output}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
